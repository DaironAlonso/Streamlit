"""
=============================================================================
SOLICITUDES DE OUTSOURCING - ARL BOLÍVAR
App Streamlit — Sin st.form, dinámico + Excel + PDF cédula por correo
=============================================================================
"""

import streamlit as st
import holidays
import datetime
import smtplib
import os
import io
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import column_index_from_string as col_idx, get_column_letter
import urllib.request
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# 1. CONFIGURACIÓN GENERAL
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Solicitudes Outsourcing — ARL Bolívar",
    page_icon="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABwAAAAcCAMAAABF0y+mAAAAvVBMVEX////5/Pz//ff/9+Lp7d2Nt5yFr4CWuY2Wu5PK2Lv/89P/zDf/z0TzyUNskjTqxkX/01HWwEqYp0O2rzv/zj7/+u3/89T/0FD/0VDuyk57mkPhxE2vsEjqyE7yy0/AtkqeqUdSiT/1zE/QvEtikUGorEewy7alqDm8tElak2alxrbe5tGaozjN2sNxpoy82MzUuz67tUm1yq2Ru6dUkl/e6+bBtTykw6VHhDWctH7h4ryJnkSMnzf/9Nnw7tcpZ/tiAAAA8UlEQVR4Ac1RQwJDQQyt+ZEyGde2ff9jdZaf6/YNY2V+jmw2m8DN5S0KxVK5Us1bhIW1uuO4ng+NZst1nHYnJOw6SIwLgoZUhE5cCKwJCJAkBG16fUk0kAQQFZrmkInRmLDBTHMSTmg6IzGABh8qGmgxXwSFy9WgJ4jUmBEiwnoTFG53xmiC/VCgBRyOIcsd7hUO+X4vyApP51BCFyH0fm9AK0K4hLxmrhdEgsEQcYjEboVQyDsgAhMagQ9oNAzVeXw4aDuHBILDc+aEhOdH2/FfjmMvf9WKNj73fn8+7/e78C6v7ZMy7W2oiDDO2cw/4wuxWRxGbur31QAAAABJRU5ErkJggg==",
    layout="wide"
)

EMAIL_USER = "notificaciones.bi.adecco@gmail.com"
EMAIL_PASS = "bgiu ydmq derj ikns"
LOGO_URL   = "https://soporte.arlsegurosbolivar.com/media/arl/arl.png"

PLANTILLA_EXCEL_PATH = os.path.join(os.path.dirname(__file__), "FORMATO.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# 2. LISTAS Y CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────

directores_sectoriales = [
    "Seleccionar Director Sectorial",
    "CAMILO ANDRES RAMIREZ OJEDA",
    "CECILIA RIOS ARIAS",
    "CRISTHIAN ANDRES RICAURTE VASQUEZ",
    "DANIEL MAURICIO JIMENEZ",
    "DANIEL MAURICIO RODRIGUEZ ALVAREZ",
    "FABIO ALBERTO ARIAS LUNA",
    "GERMAN ANTONIO NUÑEZ FLOREZ",
    "JHON JAIRO VASQUEZ",
    "JUDI ANDREA HURTADO MARTINEZ",
    "LUZ KARIME CHAVARRO CARVAJAL",
    "SILVIA CRISTINA REBOLLEDO MUNOZ",
    "YENNY EIIZABETH VERA ROMERO",
]

PROFESIONES = [
    "TECNÓLOGO SST", "TECNÓLOGO SST- AUXILIAR DE ENFERMERIA", "PROFESIONAL SST",
    "PROFESIONAL SST- AUXILIAR DE ENFERMERIA", "PROFESIONAL ESPECIALISTA", "INGENIERO ESPECIALISTA"
]

EXPERIENCIA_ANIOS = [
    "Menos de 2 AÑOS", "De 2 a 5 AÑOS", "De 5 a 9 AÑOS",
    "Mayor a 10 AÑOS", "Otra"
]

GRUPO_A_NEISY = [
    "Apia", "Armenia", "Belen De Umbria", "Circasia", "Dosquebradas",
    "Manizales", "Marmato", "Pereira", "Santa Rosa De Cabal", "Viterbo",
    "Barranquilla", "Cartagena", "Chiriguana", "El Paso", "Galapa",
    "Malambo", "Monteria", "Palmar de Varela", "Puerto Colombia", "Santa Marta",
    "Santo Tomas", "Sincelejo", "Soledad", "Turbaco", "Valledupar",
    "Acacias", "Castilla La Nueva", "Guamal", "Puerto Gaitan",
    "Villavicencio", "Yopal", "Tauramena"
]

GRUPO_B_CAMILA = [
    "Amaga", "Apartado", "Bello", "Buritica", "Caldas", "Cisneros",
    "Copacabana", "El Bagre", "Envigado", "Girardota", "Guarne",
    "Itagui", "La Estrella", "La Union", "Medellin", "Remedios",
    "Rionegro", "Sabaneta", "San Pedro De Los Milagros", "Santa Barbara",
    "Santafe De Antioquia", "Segovia", "Buenaventura", "Cali", "Cartago",
    "Dagua", "Jamundi", "Palmira", "Yumbo",
]

GRUPO_C_JINETH = [
    "Aguazul", "Anapoima", "Bogota, D.C.", "Cajica", "Chia", "Cota",
    "Cucuta", "Facatativa", "Funza", "Fusagasuga", "Gachancipa", "Granada",
    "La Calera", "Los Patios", "Madrid", "Mitu", "Mosquera", "Riohacha",
    "Sibate", "Siberia", "Soacha", "Sopo", "Tenjo", "Tocancipa",
    "Usaquen", "Villanueva", "Villapinzon", "Villeta", "Zipaquira",
]

GRUPO_D_LIZETH = [
    "El Carmen de Atrato", "Popayan", "Villa Rica", "Barrancabermeja",
    "Bucaramanga", "Floridablanca", "Giron", "Los Santos", "Piedecuesta",
    "Duitama", "Sogamoso", "Tunja", "Ibague", "Neiva", "Palermo",
    "Santa Maria", "Quibdo", "Pasto", "Ipiales", "Tumaco",
    "La Union Narino", "Samaniego", "Tuquerres",
]

TODAS_LAS_CIUDADES = sorted(GRUPO_A_NEISY + GRUPO_B_CAMILA + GRUPO_C_JINETH + GRUPO_D_LIZETH)
DIAS_SEMANA = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO", "DOMINGO", "FESTIVOS"]

CORREOS_ASESORAS = {
    "Neisy Bolanos":  "arelis.bolanos@adecco.com",
    "Camila Londono": "maria.londono@adecco.com",
    "Jineth Cortes":  "jineth.cortes@adecco.com",
    "Lizeth Garzon":  "Lizeth.GarzonZ@adecco.com",
}

CC_FIJOS = ["manuel.pimentelA@adecco.com", "ingrid.bautista@adecco.com"]
DESTINATARIOS_ARCHIVO_PLANO = [
    "MichaelE.Brochero@adecco.com",
    "manuel.pimentelA@adecco.com",
]

# DESTINATARIOS_ARCHIVO_PLANO = [
#     "dairon.alonsoh@adecco.com",
# ]

# CORREOS_ASESORAS = {
#     "Neisy Bolanos":  "desarrolladorbi7@gmail.com",
#     "Camila Londono": "desarrolladorbi7@gmail.com",
#     "Jineth Cortes":  "desarrolladorbi7@gmail.com",
#     "Lizeth Garzon":  "desarrolladorbi7@gmail.com",
# }

# CC_FIJOS = ["desarrolladorbi7@gmail.com"]

CIUDADES_PRINCIPALES = [
    "Bogota, D.C.", "Medellin", "Cali", "Barranquilla",
    "Cartagena", "Bucaramanga", "Itagui"
]

CIUDADES_INTERMEDIAS = [
    "Villavicencio", "Neiva", "Ibague", "Pereira",
    "Manizales", "Armenia", "Cucuta",
    "Pasto", "Buenaventura", "Tunja"
]

TIEMPOS_RESPUESTA = {"PRINCIPAL": 5, "INTERMEDIA": 7, "ALEJADA": 9}


# ─────────────────────────────────────────────────────────────────────────────
# 3. FUNCIONES AUXILIARES
# ─────────────────────────────────────────────────────────────────────────────

def obtener_logo() -> bytes | None:
    if not LOGO_URL:
        return None
    try:
        with urllib.request.urlopen(LOGO_URL) as resp:
            return resp.read()
    except Exception:
        return None


def obtener_plantilla_excel() -> bytes | None:
    try:
        with open(PLANTILLA_EXCEL_PATH, "rb") as f:
            return f.read()
    except Exception as e:
        st.error(f"No se encontró la plantilla Excel en: {PLANTILLA_EXCEL_PATH} — {e}")
        return None


def obtener_asesora_y_clasificacion(ciudad: str):
    if ciudad in GRUPO_A_NEISY:
        asesora = "Neisy Bolanos"
    elif ciudad in GRUPO_B_CAMILA:
        asesora = "Camila Londono"
    elif ciudad in GRUPO_C_JINETH:
        asesora = "Jineth Cortes"
    elif ciudad in GRUPO_D_LIZETH:
        asesora = "Lizeth Garzon"
    else:
        asesora = "Jineth Cortes"

    if ciudad in CIUDADES_PRINCIPALES:
        clasificacion = "PRINCIPAL"
    elif ciudad in CIUDADES_INTERMEDIAS:
        clasificacion = "INTERMEDIA"
    else:
        clasificacion = "ALEJADA"

    return asesora, clasificacion, TIEMPOS_RESPUESTA[clasificacion]


def calcular_fecha_entrega(dias_habiles: int) -> str:
    hoy = datetime.date.today()
    festivos = holidays.country_holidays("CO", years=[hoy.year, hoy.year + 1])

    if datetime.datetime.now().hour >= 12:
        hoy += datetime.timedelta(days=1)
        while hoy.weekday() >= 5 or hoy in festivos:
            hoy += datetime.timedelta(days=1)

    contador = 0
    fecha = hoy
    while contador < dias_habiles:
        if fecha.weekday() < 5 and fecha not in festivos:
            contador += 1
            if contador == dias_habiles:
                break
        fecha += datetime.timedelta(days=1)

    return fecha.strftime("%d/%m/%Y")


def generar_id_solicitud() -> str:
    return datetime.datetime.now().strftime("%Y%m%d%H%M%S")


# ─────────────────────────────────────────────────────────────────────────────
# 4. HELPERS PARA CELDAS FUSIONADAS (MergedCell fix)
# ─────────────────────────────────────────────────────────────────────────────

def get_master_cell(ws, cell_ref: str):
    """
    Retorna la celda maestra (top-left) de un rango fusionado.
    Compatible con todas las versiones de openpyxl.
    """
    cell = ws[cell_ref]
    from openpyxl.cell.cell import MergedCell
    if not isinstance(cell, MergedCell):
        return cell

    for merged_range in ws.merged_cells.ranges:
        min_col = merged_range.min_col
        min_row = merged_range.min_row
        max_col = merged_range.max_col
        max_row = merged_range.max_row
        target_row = cell.row
        target_col = cell.column
        if (min_row <= target_row <= max_row and
                min_col <= target_col <= max_col):
            return ws.cell(row=min_row, column=min_col)

    return cell


def sc(ws, cell_ref: str, value):
    """set_cell — escribe en la celda maestra aunque esté fusionada."""
    get_master_cell(ws, cell_ref).value = value


# ─────────────────────────────────────────────────────────────────────────────
# 5. HELPERS DE COLUMNAS
# ─────────────────────────────────────────────────────────────────────────────

def column_letter_to_index(col: str) -> int:
    index = 0
    for char in col.upper():
        index = index * 26 + (ord(char) - ord("A")) + 1
    return index


def index_to_column_letter(index: int) -> str:
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def get_offset_column(col_letter: str, offset: int) -> str:
    return index_to_column_letter(column_letter_to_index(col_letter) + offset)


# ─────────────────────────────────────────────────────────────────────────────
# 6. DILIGENCIAR FORMATO EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def diligenciar_formato_excel(datos: dict, plantilla_bytes: bytes, logo_bytes: bytes | None) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(plantilla_bytes))
    ws = wb["FORMATO"]
    row = datos

    def fill_si_no_block(q_map: dict, displacement: int):
        for q, si_cell in q_map.items():
            response = str(row.get(q, "")).strip().upper()
            if response not in ("SI", "NO"):
                continue
            col_letter = "".join(c for c in si_cell if c.isalpha())
            row_number = "".join(c for c in si_cell if c.isdigit())
            no_col_letter = get_offset_column(col_letter, displacement)
            no_cell = no_col_letter + row_number
            if response == "SI":
                get_master_cell(ws, si_cell).value = "X"
            elif response == "NO":
                get_master_cell(ws, no_cell).value = "X"

    # ── I. Información general ─────────────────────────────────────────────
    sc(ws, "F8",  row.get("Q2", ""))
    if row.get("Q6") == "NUEVO":
        sc(ws, "Z8",  "X")
    elif row.get("Q6") == "REEMPLAZO":
        sc(ws, "AE8", "X")

    sc(ws, "P10",  row.get("Q7",  ""))
    sc(ws, "Q12",  row.get("Q8",  ""))
    sc(ws, "AG12", row.get("Q9",  ""))
    sc(ws, "M14",  row.get("Q10", ""))
    sc(ws, "H16",  row.get("Q11", ""))
    sc(ws, "AF16", row.get("Q12", ""))
    sc(ws, "G18",  row.get("Q13", ""))
    sc(ws, "X18",  row.get("Q14", ""))

    # ── II. Cargo ──────────────────────────────────────────────────────────
    sc(ws, "G23", row.get("Q15", ""))
    sc(ws, "Q29", row.get("Q20", ""))
    sc(ws, "G35", row.get("Q21", 0))

    if row.get("Q22") == "FIJO":
        sc(ws, "AG35", "X")
    elif row.get("Q22") == "INTERDISCIPLINARIO":
        sc(ws, "Z35",  "X")

    if row.get("Q23") == "150 HORAS":
        sc(ws, "R37", "X")
    elif row.get("Q23") == "75 HORAS":
        sc(ws, "L37", "X")

    sc(ws, "Z37", row.get("Q24", ""))
    sc(ws, "O39", row.get("Q25", ""))

    dias_raw = str(row.get("Q26", ""))
    dias_seleccionados = [d.strip().upper() for d in dias_raw.split(";") if d.strip()]
    casillas_dias = {
        "LUNES": "H41", "MARTES": "K41", "MIERCOLES": "N41", "JUEVES": "Q41",
        "VIERNES": "T41", "SABADO": "W41", "DOMINGO": "Z41", "FESTIVOS": "AG41"
    }
    for dia, celda in casillas_dias.items():
        get_master_cell(ws, celda).value = "X" if dia in dias_seleccionados else ""

    if row.get("Q22") == "INTERDISCIPLINARIO":
        agr_data = row.get("agr_data", [])
        celdas_agr = [
            ("J49", "AG49"),
            ("J51", "AG51"),
            ("J53", "AG53"),
            ("J55", "AG55"),
            ("J57", "AG57"),
        ]
        for idx, agr in enumerate(agr_data):
            if idx < len(celdas_agr):
                celda_nombre, celda_horas = celdas_agr[idx]
                sc(ws, celda_nombre, agr.get("agr", ""))
                sc(ws, celda_horas,  agr.get("horas", 0))

    sc(ws, "G43", row.get("Q27", ""))

    opciones_riesgo = {1: "H45", 2: "K45", 3: "N45", 4: "Q45", 5: "T45"}
    try:
        riesgo = int(row.get("Q28", 0))
        if riesgo in opciones_riesgo:
            get_master_cell(ws, opciones_riesgo[riesgo]).value = "X"
    except Exception:
        pass

    sc(ws, "AA45", row.get("Q29", ""))

    if row.get("Q30") == "MOTO":
        sc(ws, "R59", "X")
    elif row.get("Q30") == "VEHICULO":
        sc(ws, "W59", "X")
    sc(ws, "AC59", row.get("Q31", ""))

    # ── III. Auxilios ──────────────────────────────────────────────────────
    def x(q):  return "X" if str(row.get(q, "")).upper() == "SI" else ""
    def no(q): return "X" if str(row.get(q, "")).upper() == "NO" else ""

    get_master_cell(ws, "I62").value  = x("Q32");  get_master_cell(ws, "L62").value  = no("Q32")
    get_master_cell(ws, "Q62").value  = row.get("Q33", "")
    get_master_cell(ws, "AC62").value = row.get("Q34", "")

    get_master_cell(ws, "I63").value  = x("Q35");  get_master_cell(ws, "L63").value  = no("Q35")
    get_master_cell(ws, "Q63").value  = row.get("Q36", "")
    get_master_cell(ws, "AC63").value = row.get("Q37", "")

    get_master_cell(ws, "I65").value  = x("Q38");  get_master_cell(ws, "L65").value  = no("Q38")
    get_master_cell(ws, "Q65").value  = row.get("Q39", "")
    get_master_cell(ws, "AC65").value = row.get("Q40", "")

    get_master_cell(ws, "H67").value  = row.get("Q42_texto", "")
    get_master_cell(ws, "Q67").value  = row.get("Q43_frec",  "")
    get_master_cell(ws, "AC67").value = row.get("Q44_valor", "")

    # ── IV. Competencias ───────────────────────────────────────────────────
    comp_map = {
        "Q56": "O74", "Q57": "O76", "Q58": "O78",
        "Q59": "O80", "Q60": "O82", "Q61": "O84",
        "Q62": "O86", "Q63": "O88", "Q64": "O90",
        "Q65": "O92",
        "Q66": "AG74", "Q67": "AG76", "Q68": "AG78",
    }
    fill_si_no_block(comp_map, 3)
    get_master_cell(ws, "S84").value = row.get("Q71_extra", "")

    # ── V. EPPs / Dotación ─────────────────────────────────────────────────
    epps_map = {
        "Q72": "O98",  "Q73": "O100", "Q74": "O102", "Q75": "O104",
        "Q76": "O106", "Q77": "O108", "Q78": "O110",
        "Q79": "AG98", "Q80": "AG100","Q81": "AG102","Q82": "AG104",
        "Q83": "AG106","Q84": "AG108",
    }
    fill_si_no_block(epps_map, 3)
    get_master_cell(ws, "U111").value = row.get("Q86_texto", "")

    extras_map = {
        "Q87": "O114", "Q88": "O116",
        "Q89": "AG114","Q90": "AG116",
        "Q70": "AG80",
        "Q86": "AG110",
        "Q92": "AG118"
    }
    fill_si_no_block(extras_map, 3)

    # ── VI. Recomendaciones ────────────────────────────────────────────────
    get_master_cell(ws, "A122").value = row.get("Q91", "")
    get_master_cell(ws, "Z119").value = row.get("Q93", "")

    # ── Logo ───────────────────────────────────────────────────────────────
    if logo_bytes:
        try:
            img = XLImage(io.BytesIO(logo_bytes))
            img.width  = 120
            img.height = 50
            img.anchor = "B2"
            ws.add_image(img)
        except Exception as e:
            st.warning(f"No se pudo insertar el logo en el Excel: {e}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ─────────────────────────────────────────────────────────────────────────────
# 7. ENVÍO DE CORREO
# ─────────────────────────────────────────────────────────────────────────────

def enviar_correo(datos: dict, xlsx_bytes: bytes, archivos_hv: list[dict]) -> bool:
    ciudad = str(datos.get("Q25", "")).strip()
    asesora, clasificacion, dias_hab = obtener_asesora_y_clasificacion(ciudad)
    fecha_entrega = calcular_fecha_entrega(dias_hab)

    correo_agr  = datos.get("Q11", "")
    id_sol      = datos.get("id_solicitud", "N/A")
    perfil      = datos.get("Q15", "")
    nombre_base = f"{id_sol} - {perfil} en {ciudad}"

    cuerpo = f"""
    <html><body>
    <p>Estimada Asesora,</p>
    <p>Se adjunta el formulario de Solicitud de Outsourcing para el perfil:</p>
    <p><b>{perfil}</b> en <b>{ciudad}</b>.</p>

    <p><b>NOTA IMPORTANTE:</b> Según la clasificación de la ciudad (<b>{clasificacion}</b>),
    el tiempo de respuesta es de <b>{dias_hab} días hábiles</b>.</p>
    <p>Fecha tentativa de entrega: <b>{fecha_entrega}</b>.</p>
    <p>Número de vacantes solicitadas: <b>{datos.get("Q24", "")}</b>.</p>

    <p>Por favor, revisarlo y dar continuidad al proceso.</p>
    <br>
    <p><b>Información del Solicitante:</b></p>
    <ul>
      <li><b>AGR:</b> {datos.get("Q10","")}</li>
      <li><b>Correo AGR:</b> {correo_agr}</li>
      <li><b>Tipo de Asignación:</b> {datos.get("Q22","")}</li>
      <li><b>Empresa Afiliada:</b> {datos.get("Q8","")}</li>
      <li><b>NIT:</b> {datos.get("Q9","")}</li>
    </ul>
    <br>
    <p>Atentamente,<br>Equipo de Notificaciones BI — Adecco.</p>
    </body></html>
    """

    destinatario = CORREOS_ASESORAS.get(asesora, "jineth.cortes@adecco.com")
    cc_lista     = CC_FIJOS + ([correo_agr] if correo_agr else [])

    msg = MIMEMultipart()
    msg["From"]    = EMAIL_USER
    msg["To"]      = destinatario
    msg["Cc"]      = "; ".join(cc_lista)
    msg["Subject"] = f"ID {id_sol} - NUEVA SOLICITUD DE OUTSOURCING: {perfil} en {ciudad}"
    msg.attach(MIMEText(cuerpo, "html"))

    parte_xlsx = MIMEBase("application", "octet-stream")
    parte_xlsx.set_payload(xlsx_bytes)
    encoders.encode_base64(parte_xlsx)
    parte_xlsx.add_header(
        "Content-Disposition", "attachment",
        filename=("utf-8", "", f"{nombre_base}.xlsx")
    )
    msg.attach(parte_xlsx)

    for archivo in archivos_hv:
        parte_hv = MIMEBase("application", "octet-stream")
        parte_hv.set_payload(archivo["bytes"])
        encoders.encode_base64(parte_hv)
        parte_hv.add_header(
            "Content-Disposition", "attachment",
            filename=("utf-8", "", f"{archivo['nombre']}")
        )
        msg.attach(parte_hv)

    try:
        todos = [destinatario] + cc_lista
        with smtplib.SMTP("smtp.gmail.com", 587) as servidor:
            servidor.starttls()
            servidor.login(EMAIL_USER, EMAIL_PASS)
            servidor.sendmail(EMAIL_USER, todos, msg.as_string())
        return True
    except Exception as e:
        st.error(f"Error al enviar correo: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# 8. GENERAR ARCHIVO PLANO — TODOS LOS CAMPOS (incluyendo ocultos/condicionales)
# ─────────────────────────────────────────────────────────────────────────────

def generar_excel_plano(datos: dict) -> tuple[str, bytes]:
    id_sol   = str(datos.get("id_solicitud", "N/A"))
    agr_data = datos.get("agr_data", [])

    # ── AGR: siempre 5 slots fijos (vacíos si no aplica) ──────────────────
    # Garantiza estructura uniforme independientemente de si es FIJO o INTERDISCIPLINARIO
    agr_cols = {}
    for i in range(1, 6):
        agr = agr_data[i - 1] if i <= len(agr_data) else {}
        agr_cols[f"AGR_{i}_NOMBRE"]          = agr.get("agr",   "")
        agr_cols[f"AGR_{i}_HORAS_MENSUALES"] = agr.get("horas", "")

    # ── Campos calculados ─────────────────────────────────────────────────
    ciudad = str(datos.get("Q25", "")).strip()
    try:
        asesora_asignada, clasificacion_ciudad, dias_respuesta = obtener_asesora_y_clasificacion(ciudad)
        fecha_entrega_estimada = calcular_fecha_entrega(dias_respuesta)
    except Exception:
        asesora_asignada        = ""
        clasificacion_ciudad    = ""
        dias_respuesta          = ""
        fecha_entrega_estimada  = ""

    # ── Columnas del archivo plano ─────────────────────────────────────────
    # TODOS los campos del formulario aparecen aquí, incluyendo los condicionales/ocultos.
    # Los campos que no aplican en un registro llevarán cadena vacía o 0.
    columnas = {
        # ── Metadatos de la solicitud ──────────────────────────────────────
        "ID_SOLICITUD":                        id_sol,
        "FECHA":                               datos.get("Q2",  ""),
        "ASESORA_ASIGNADA":                    asesora_asignada,
        "CLASIFICACION_CIUDAD":                clasificacion_ciudad,
        "DIAS_HABILES_RESPUESTA":              dias_respuesta,
        "FECHA_ENTREGA_ESTIMADA":              fecha_entrega_estimada,

        # ── I. Información general ─────────────────────────────────────────
        "TIPO_SOLICITUD":                      datos.get("Q6",  ""),
        # Condicional: solo se llena si Q6 == "REEMPLAZO", pero siempre aparece en el plano
        "TRABAJADOR_REEMPLAZAR":               datos.get("Q7",  ""),
        "RAZON_SOCIAL_EMPRESA":                datos.get("Q8",  ""),
        "NIT_EMPRESA":                         datos.get("Q9",  ""),
        "AGR_SOLICITANTE":                     datos.get("Q10", ""),
        "CORREO_AGR":                          datos.get("Q11", ""),
        "CELULAR_AGR":                         datos.get("Q12", ""),
        "DIRECCION_SECTORIAL":                 datos.get("Q13", ""),
        "DIRECTOR_SECTORIAL":                  datos.get("Q14", ""),

        # ── II. Cargo ──────────────────────────────────────────────────────
        # Q15 ya contiene la concatenación completa incluyendo sub-selectores ocultos
        # (profesional_sst / profesional_sst_ae / otro_profesional / ingeniero_especialista / otro_ingeniero)
        "PERFIL":                              datos.get("Q15", ""),
        "EXPERIENCIA_REQUERIDA":               datos.get("Q20", ""),
        "SALARIO_FUERA_TABLA":                 datos.get("Q21", ""),
        "TIPO_ASIGNACION":                     datos.get("Q22", ""),
        # AGR slots fijos 1-5 (vacíos si Q22 == "FIJO")
        **agr_cols,
        "TIEMPO_SERVICIO":                     datos.get("Q23", ""),
        "NUMERO_VACANTES":                     datos.get("Q24", ""),
        "CIUDAD_SERVICIO":                     datos.get("Q25", ""),
        "DIAS_SERVICIO":                       datos.get("Q26", ""),
        "HORARIO_SERVICIO":                    datos.get("Q27", ""),
        "CLASE_RIESGO":                        datos.get("Q28", ""),
        "SECTOR_ECONOMICO":                    datos.get("Q29", ""),
        "TRANSPORTE_PROPIO":                   datos.get("Q30", ""),
        # Condicional: solo se llena si Q30 != "NINGUNO", pero siempre aparece en el plano
        "AUXILIO_TRANSPORTE_PROPIO":           datos.get("Q31", ""),

        # ── III. Auxilios Autorizados ──────────────────────────────────────
        "TRANSPORTE_URBANO":                   datos.get("Q32", ""),
        # Condicionales de Q32: solo se llenan si Q32 == "SI"
        "FRECUENCIA_TRANSPORTE_URBANO":        datos.get("Q33", ""),
        "VALOR_TRANSPORTE_URBANO":             datos.get("Q34", ""),

        "TRANSPORTE_INTERMUNICIPAL":           datos.get("Q35", ""),
        # Condicionales de Q35: solo se llenan si Q35 == "SI"
        "FRECUENCIA_TRANSPORTE_INTERMUNICIPAL": datos.get("Q36", ""),
        "VALOR_TRANSPORTE_INTERMUNICIPAL":     datos.get("Q37", ""),

        "COMUNICACION":                        datos.get("Q38", ""),
        # Condicionales de Q38: solo se llenan si Q38 == "SI"
        "FRECUENCIA_COMUNICACION":             datos.get("Q39", ""),
        "VALOR_COMUNICACION":                  datos.get("Q40", ""),

        "OTRO_AUXILIO":                        datos.get("Q41", ""),
        # Condicionales de Q41: solo se llenan si Q41 == "SI"
        "OTRO_AUXILIO_CUAL":                   datos.get("Q42_texto", ""),
        "OTRO_AUXILIO_FRECUENCIA":             datos.get("Q43_frec",  ""),
        "OTRO_AUXILIO_VALOR":                  datos.get("Q44_valor", ""),

        # ── IV. Competencias Técnicas ──────────────────────────────────────
        "PRUEBA_TECNICA_PROFESION_BASE":       datos.get("Q56", ""),
        "PRUEBA_OFIMATICA":                    datos.get("Q57", ""),
        "PRUEBA_TECNICA_SIG":                  datos.get("Q58", ""),

        # ── IV. Competencias Blandas ───────────────────────────────────────
        "ORIENTACION_RESULTADOS":              datos.get("Q59", ""),
        "ORIENTACION_CLIENTE":                 datos.get("Q60", ""),
        "ANALISIS_SOLUCION_PROBLEMAS":         datos.get("Q61", ""),
        "ADAPTACION_CAMBIO":                   datos.get("Q62", ""),
        "AUTOMANEJO_AUTODESARROLLO":           datos.get("Q63", ""),
        "COMUNICACION_COMPETENCIA":            datos.get("Q64", ""),
        "TRABAJO_EN_EQUIPO":                   datos.get("Q65", ""),
        "DESARROLLO_RELACIONES":               datos.get("Q66", ""),
        "LIDERAR_EQUIPOS":                     datos.get("Q67", ""),
        "PLANIFICACION_ESTRATEGICA":           datos.get("Q68", ""),
        "OTRA_COMPETENCIA":                    datos.get("Q70", ""),
        # Condicional: solo se llena si Q70 == "SI"
        "DESCRIPCION_OTRA_COMPETENCIA":        datos.get("Q71_extra", ""),

        # ── V. EPPs ────────────────────────────────────────────────────────
        "EPP_CASCO_DIELECTRICO":               datos.get("Q72", ""),
        "EPP_CASCO_BARBUQUEJO":                datos.get("Q73", ""),
        "EPP_PROTECTOR_AUDITIVO_COPA":         datos.get("Q74", ""),
        "EPP_PROTECTOR_AUDITIVO_INSERCION":    datos.get("Q75", ""),
        "EPP_MONOGAFA":                        datos.get("Q76", ""),
        "EPP_PROTECCION_RESPIRATORIA":         datos.get("Q77", ""),
        "EPP_PROTECCION_VISUAL":               datos.get("Q78", ""),

        # ── V. Dotación ────────────────────────────────────────────────────
        "DOTACION_UNIFORME_ANTIFLUIDO":        datos.get("Q79", ""),
        "DOTACION_CHAQUETA":                   datos.get("Q80", ""),
        "DOTACION_CAMISA":                     datos.get("Q81", ""),
        "DOTACION_JEAN":                       datos.get("Q82", ""),
        "DOTACION_BOTAS_ANTIDESLIZANTE":       datos.get("Q83", ""),
        "DOTACION_BOTAS_ANTIPERFORANTE":       datos.get("Q84", ""),
        "OTRA_DOTACION":                       datos.get("Q86", ""),
        # Condicional: solo se llena si Q86 == "SI"
        "DESCRIPCION_OTRA_DOTACION":           datos.get("Q86_texto", ""),

        # ── V. Equipo de cómputo ───────────────────────────────────────────
        "EQUIPO_COMPUTO_BASICO":               datos.get("Q87", ""),
        "EQUIPO_COMPUTO_MAYOR_CAPACIDAD":      datos.get("Q88", ""),

        # ── V. Cursos especiales ───────────────────────────────────────────
        "CURSO_ALTURAS":                       datos.get("Q89", ""),
        "CURSO_ESPACIOS_CONFINADOS":           datos.get("Q90", ""),

        # ── VI. Exámenes especializados ────────────────────────────────────
        "REQUIERE_EXAMENES_INGRESO":           datos.get("Q92", ""),
        # Condicional: solo se llena si Q92 == "SI"
        "EXAMENES_CUALES":                     datos.get("Q93", ""),

        # ── VII. Recomendaciones ───────────────────────────────────────────
        "RECOMENDACIONES_GENERALES":           datos.get("Q91", ""),
    }

    # ── Crear Excel ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitud"

    font_header = Font(bold=True, color="FFFFFF", size=10)
    fill_header = PatternFill("solid", fgColor="016D38")
    font_valor  = Font(size=10)
    fill_valor  = PatternFill("solid", fgColor="F9F9F9")
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    align_header = Alignment(horizontal="center", vertical="center",
                             wrap_text=True, text_rotation=90)
    align_valor  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, (header, valor) in enumerate(columnas.items(), start=1):
        ch = ws.cell(row=1, column=col_num, value=header)
        ch.font      = font_header
        ch.fill      = fill_header
        ch.alignment = align_header
        ch.border    = borde

        cv = ws.cell(row=2, column=col_num, value=str(valor) if valor is not None else "")
        cv.font      = font_valor
        cv.fill      = fill_valor
        cv.alignment = align_valor
        cv.border    = borde

        ws.column_dimensions[get_column_letter(col_num)].width = 18

    ws.row_dimensions[1].height = 80
    ws.row_dimensions[2].height = 25
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f"{id_sol}_Solicitud plano.xlsx"
    return nombre_archivo, output.read()


# ─────────────────────────────────────────────────────────────────────────────
# 9. ENVÍO ARCHIVO PLANO
# ─────────────────────────────────────────────────────────────────────────────

def enviar_archivo_plano(datos: dict) -> bool:
    nombre_archivo, excel_bytes = generar_excel_plano(datos)
    id_sol = str(datos.get("id_solicitud", "N/A"))
    ciudad = str(datos.get("Q25", "")).strip()
    perfil = str(datos.get("Q15", "")).strip()

    msg = MIMEMultipart()
    msg["From"]    = EMAIL_USER
    msg["To"]      = "; ".join(DESTINATARIOS_ARCHIVO_PLANO)
    msg["Subject"] = f"Archivo plano solicitud outsourcing"

    cuerpo = (
        f"Se adjunta el resumen en Excel de la solicitud de outsourcing"
    )
    msg.attach(MIMEText(cuerpo, "plain", "utf-8"))

    parte_xlsx = MIMEBase("application", "octet-stream")
    parte_xlsx.set_payload(excel_bytes)
    encoders.encode_base64(parte_xlsx)
    parte_xlsx.add_header(
        "Content-Disposition", "attachment",
        filename=("utf-8", "", nombre_archivo)
    )
    msg.attach(parte_xlsx)

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as servidor:
            servidor.starttls()
            servidor.login(EMAIL_USER, EMAIL_PASS)
            servidor.sendmail(EMAIL_USER, DESTINATARIOS_ARCHIVO_PLANO, msg.as_string())
        return True
    except Exception as e:
        st.error(f"Error al enviar archivo plano: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# 10. INTERFAZ STREAMLIT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    st.markdown("""
    <style>
    .section-header {
        background: #016d38;
        color: white;
        padding: 8px 14px;
        border-radius: 6px;
        font-weight: 600;
        font-size: 0.95rem;
        margin: 20px 0 10px 0;
    }
    .info-box {
        background: #d4e6f1;
        border-left: 4px solid #016d38;
        padding: 10px 14px;
        border-radius: 4px;
        font-size: 0.88rem;
        margin: 8px 0;
    }
    </style>
    """, unsafe_allow_html=True)

    col_logo, col_titulo = st.columns([1, 4])
    with col_logo:
        logo_bytes = obtener_logo()
        if logo_bytes:
            st.image(logo_bytes, width=300)
    with col_titulo:
        st.markdown("## Solicitud de Outsourcing de Servicios Especializados de Gestion")
        st.markdown("**ARL Seguros Bolivar** — Complete el formulario y presione **Enviar Solicitud**.")

    st.divider()

    # ── I. Información general ─────────────────────────────────────────────
    st.markdown('<div class="section-header">I. INFORMACION GENERAL DE SOLICITUD</div>',
                unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        Q6 = st.selectbox("Tipo de solicitud *", ["NUEVO", "REEMPLAZO"])
    with c2:
        if Q6 == "REEMPLAZO":
            Q7 = st.text_input("Nombre del trabajador a reemplazar *")
        else:
            Q7 = ""

    c1, c2 = st.columns(2)
    with c1:
        Q8 = st.text_input("Razon social empresa afiliada *")
    with c2:
        Q9 = st.text_input("NIT de la empresa *")

    c1, c2, c3 = st.columns(3)
    with c1:
        Q10 = st.text_input("AGR Solicitante (Lider o Responsable) *")
    with c2:
        Q11 = st.text_input("Correo electronico AGR *")
    with c3:
        Q12 = st.text_input("Numero de celular AGR *")

    c1, c2 = st.columns(2)
    with c1:
        Q13 = st.text_input("Direccion Sectorial *")
    with c2:
        Q14 = st.selectbox("Nombre Director Sectorial *", options=directores_sectoriales)

    # ── II. Información del cargo ──────────────────────────────────────────
    st.markdown('<div class="section-header">II. INFORMACION GENERAL DEL CARGO</div>',
                unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        Q15 = st.selectbox("Profesion / Perfil principal *", PROFESIONES)
    with c2:
        if Q15 == "PROFESIONAL SST":
            profesional_sst = st.selectbox("Profesional SST *", [
                "ADMINISTRADOR SST",
                "INGENIERO EN SEGURIDAD Y SALUD EN EL TRABAJO",
                "PROFESIONAL SST"
            ])
            Q15 = Q15 + " - " + profesional_sst
        elif Q15 == "PROFESIONAL ESPECIALISTA":
            profesional_sst_ae = st.selectbox("Profesional Especialista *", [
                "ADMINISTRADOR AMBIENTAL", "IADMINISTRADOR DE EMPRESAS", "ENFERMERO",
                "FISIOTERAPEUTA", "FONOAUDIÓLOGO",
                "INGENIERO EN SEGURIDAD Y SALUD EN EL TRABAJO",
                "MÉDICO", "PSICÓLOGO", "TERAPEUTA OCUPACIONAL", "OTRO"
            ])
            Q15 = Q15 + " - " + profesional_sst_ae
            if profesional_sst_ae == "OTRO":
                otro_profesional = st.text_input("Indique que otra profesión requiere *", value="")
                Q15 = Q15 + " - " + otro_profesional
        elif Q15 == "INGENIERO ESPECIALISTA":
            ingeniero_especialista = st.selectbox("Ingeniero Especialista *", [
                "INGENIERO AMBIENTAL", "INGENIERO AMBIENTAL Y SANITARIO",
                "INGENIERO DE PROCESOS", "INGENIERO DE PRODUCCIÓN",
                "INGENIERO ELECTRICISTA", "INGENIERO ELECTROMECÁNICO",
                "INGENIERO INDUSTRIAL", "INGENIERO MECÁNICO",
                "INGENIERO QUÍMICO", "INGENIERO SANITARIO", "OTRO",
            ])
            Q15 = Q15 + " - " + ingeniero_especialista
            if ingeniero_especialista == "OTRO":
                otro_ingeniero = st.text_input("Indique que otro ingeniero requiere *", value="")
                Q15 = Q15 + " - " + otro_ingeniero

    c1, c2 = st.columns(2)
    with c1:
        Q20 = st.selectbox("Experiencia requerida (anos minimos) *", EXPERIENCIA_ANIOS)
    with c2:
        if Q20 == "Otra":
            Q20 = st.text_input("Mencione experiencia diferente a lo anterior")

    Q21_raw = st.text_input(
        "En caso de que el salario no esté en la tabla autorizada, indicar el valor (si está en tabla colocar 0)",
        value="0"
    )
    try:
        Q21 = int(Q21_raw.replace(".", "").replace(",", "").replace("$", "").strip())
        st.caption(f"💰 Valor: **${Q21:,.0f}**".replace(",", "X").replace(".", ",").replace("X", "."))
    except ValueError:
        st.error("Por favor ingresa un valor numérico válido.")
        Q21 = 0

    c1, c2 = st.columns(2)
    with c1:
        Q22 = st.selectbox("Asignación *", ["INTERDISCIPLINARIO", "FIJO"])
    with c2:
        if Q22 == "INTERDISCIPLINARIO":
            if "num_agr" not in st.session_state:
                st.session_state.num_agr = 1
            agr_data = []
            for i in range(1, st.session_state.num_agr + 1):
                c1, c2, c3 = st.columns(3)
                with c1:
                    nombre = st.text_input(f"AGR {i}", key=f"agr_nombre_{i}")
                with c2:
                    horas = st.number_input(
                        f"Horas mensuales (AGR {i})",
                        min_value=0, step=1, value=0,
                        key=f"agr_horas_{i}"
                    )
                with c3:
                    if i == st.session_state.num_agr and i < 5:
                        agregar = st.selectbox(
                            "Adicionar otro AGR",
                            options=["No", "SI"],
                            key=f"agr_add_{i}"
                        )
                        if agregar == "SI":
                            st.session_state.num_agr += 1
                            st.rerun()
                agr_data.append({"agr": nombre, "horas": horas})
            if st.session_state.num_agr > 1:
                if st.button("➖ Quitar último AGR"):
                    st.session_state.num_agr -= 1
                    st.rerun()
        else:
            st.session_state.num_agr = 1
            agr_data = []

    c1, c2, c3 = st.columns(3)
    with c1:
        Q23 = st.selectbox("Tiempo de prestacion del servicio *", ["75 HORAS", "150 HORAS"])
    with c2:
        Q24 = st.number_input("Numero de vacantes *", min_value=1, step=1, value=1)
    with c3:
        Q25 = st.selectbox("Ciudad/Municipio donde se prestara el servicio *", TODAS_LAS_CIUDADES)

    c1, c2 = st.columns(2)
    with c1:
        Q26 = st.multiselect("Dias de servicio *", DIAS_SEMANA)
    with c2:
        Q27 = st.text_input("Horario de servicio *")

    c1, c2 = st.columns(2)
    with c1:
        Q28 = st.selectbox("Clase de riesgo *", [1, 2, 3, 4, 5])
    with c2:
        Q29 = st.text_input("Sector economico *")

    c1, c2 = st.columns(2)
    with c1:
        Q30 = st.selectbox(
            "Requiere algún tipo de transporte propio *",
            ["NINGUNO", "MOTO", "VEHICULO"]
        )
    with c2:
        if Q30 != "NINGUNO":
            Q31 = st.number_input("Auxilio", min_value=0, step=10000, value=0)
        else:
            Q31 = 0

    # ── III. Auxilios Autorizados ──────────────────────────────────────────
    st.markdown('<div class="section-header">III. AUXILIOS AUTORIZADOS</div>',
                unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1: Q32 = st.selectbox("Transporte Urbano", ["NO", "SI"])
    if Q32 == "SI":
        with c2: Q33 = st.selectbox("Frecuencia T. Urbano", ["", "QUINCENAL", "MENSUAL", "ANUAL"])
        with c3: Q34 = st.number_input("Valor T. Urbano ($)", min_value=0, step=10000, value=0)
    else:
        Q33 = ""
        Q34 = 0

    c1, c2, c3 = st.columns(3)
    with c1: Q35 = st.selectbox("Transporte Intermunicipal", ["NO", "SI"])
    if Q35 == "SI":
        with c2: Q36 = st.selectbox("Frecuencia T. Intermunicipal", ["", "QUINCENAL", "MENSUAL", "ANUAL"])
        with c3: Q37 = st.number_input("Valor T. Intermunicipal ($)", min_value=0, step=10000, value=0)
    else:
        Q36 = ""
        Q37 = 0

    c1, c2, c3 = st.columns(3)
    with c1: Q38 = st.selectbox("Comunicacion", ["NO", "SI"])
    if Q38 == "SI":
        with c2: Q39 = st.selectbox("Frecuencia Comunicacion", ["", "QUINCENAL", "MENSUAL", "ANUAL"])
        with c3: Q40 = st.number_input("Valor Comunicacion ($)", min_value=0, step=10000, value=0)
    else:
        Q39 = ""
        Q40 = 0

    c1, c2, c3 = st.columns(3)
    with c1: Q41 = st.selectbox("Otro auxilio", ["NO", "SI"])
    if Q41 == "SI":
        with c2: Q42_texto = st.text_input("Cual otro auxilio?")
        with c3: Q43_frec = st.selectbox("Frecuencia otro auxilio", ["", "QUINCENAL", "MENSUAL", "ANUAL"])
        Q44_valor = st.number_input("Valor otro auxilio ($)", min_value=0, step=10000, value=0)
    else:
        Q42_texto = ""
        Q43_frec  = ""
        Q44_valor = 0

    # ── IV. Competencias ───────────────────────────────────────────────────
    st.markdown('<div class="section-header">IV. COMPETENCIAS TECNICAS REQUERIDAS</div>',
                unsafe_allow_html=True)

    st.markdown("**COMPETENCIAS BLANDAS**")
    comp_labels_form = [
        ("Q59", "Orientación a resultados"),
        ("Q60", "Orientación al cliente"),
        ("Q61", "Análisis y solución de problemas"),
        ("Q62", "Adaptación al cambio"),
        ("Q63", "Automanejo y autodesarrollo"),
        ("Q64", "Comunicación"),
        ("Q65", "Trabajo en equipo"),
        ("Q66", "Desarrollo de relaciones"),
        ("Q67", "Liderar equipos"),
        ("Q68", "Planificación estratégica y capacidad de decisión"),
    ]
    comp_vals = {}
    for i in range(0, len(comp_labels_form), 3):
        cols = st.columns(3)
        for j, (qkey, qlabel) in enumerate(comp_labels_form[i:i+3]):
            with cols[j]:
                comp_vals[qkey] = st.selectbox(qlabel, ["", "NO", "SI"], key=f"comp_{qkey}")

    st.markdown("**COMPETENCIAS TÉCNICAS**")
    c1, c2, c3 = st.columns(3)
    with c1:
        Q56 = st.selectbox("Prueba técnica especifica de acuerdo a profesión base", ["", "NO", "SI"])
    with c2:
        Q57 = st.selectbox("Prueba ofimática (mínimo Excel intermedio)", ["", "NO", "SI"])
    with c3:
        Q58 = st.selectbox("Prueba técnica especifica SIG", ["", "NO", "SI"])

    c1, c2 = st.columns(2)
    with c1:
        Q70 = st.selectbox("Otra", ["NO", "SI"])
    with c2:
        if Q70 == "SI":
            Q71_extra = st.text_input("Descripcion adicional competencias tecnicas especificas")
        else:
            Q71_extra = ""

    # ── V. Dotación / EPPs / Equipos / Cursos ─────────────────────────────
    st.markdown("<div class=\"section-header\">V. DOTACIÓN - EPP'S - EQUIPO DE COMPUTO - CURSOS ADICIONALES</div>",
                unsafe_allow_html=True)

    st.markdown("**Elementos de protección personal**")
    epp_items = [
        ("Q72", "Casco dieléctrico"),
        ("Q73", "Casco dieléctrico con barbuquejo"),
        ("Q74", "Protector auditivo de copa"),
        ("Q75", "Protector auditivo de inserción"),
        ("Q76", "Monogafa antiempañante con filtro UV quimicoresistente"),
        ("Q77", "Protección respiratoria para material particulado, vapores orgánicos, vapores ácidos, humos metálicos"),
        ("Q78", "Protección visual (lente claro - oscuro)"),
    ]
    epp_vals = {}
    for i in range(0, len(epp_items), 3):
        cols = st.columns(3)
        for j, (qk, ql) in enumerate(epp_items[i:i+3]):
            with cols[j]:
                epp_vals[qk] = st.selectbox(ql, ["NO", "SI"], key=f"epp_{qk}")

    st.markdown("**Dotacion**")
    dot_items = [
        ("Q79", "Uniforme anti fluido"),
        ("Q80", "Chaqueta"),
        ("Q81", "Camisa"),
        ("Q82", "Jean"),
        ("Q83", "Botas dieléctricas / antideslizante"),
        ("Q84", "Botas dieléctricas / antiperforante"),
    ]
    dot_vals = {}
    for i in range(0, len(dot_items), 3):
        cols = st.columns(3)
        for j, (qk, ql) in enumerate(dot_items[i:i+3]):
            with cols[j]:
                dot_vals[qk] = st.selectbox(ql, ["NO", "SI"], key=f"dot_{qk}")

    c1, c2 = st.columns(2)
    with c1:
        Q86 = st.selectbox("¿Otro?", ["NO", "SI"])
    with c2:
        if Q86 == "SI":
            Q86_texto = st.text_input("Cual otro elemento de dotación se requiere?")
        else:
            Q86_texto = ""

    st.markdown("**Equipo de computo**")
    c1, c2 = st.columns(2)
    with c1: Q87 = st.selectbox("Equipo de computo basico", ["NO", "SI"])
    with c2: Q88 = st.selectbox("Equipo de computo mayor capacidad", ["NO", "SI"])

    st.markdown("**Cursos especiales**")
    c1, c2 = st.columns(2)
    with c1: Q89 = st.selectbox("Curso trabajo seguro en alturas", ["NO", "SI"])
    with c2: Q90 = st.selectbox("Curso espacios confinados", ["NO", "SI"])

    # ── VI. Exámenes especializados ────────────────────────────────────────
    st.markdown('<div class="section-header">VI. EXAMANES ESPECIALIZADOS</div>',
                unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1: Q92 = st.selectbox("¿El profesional seleccionado requiere examenes de ingreso especializados?", ["NO", "SI"])
    with c2:
        if Q92 == "SI":
            Q93 = st.text_input("¿Cuales?")
        else:
            Q93 = ""

    # ── VII. Recomendaciones ────────────────────────────────────────────────
    st.markdown('<div class="section-header">VII. RECOMENDACIONES PARA TENER EN CUENTA DURANTE EL PROCESO</div>',
                unsafe_allow_html=True)
    Q91 = st.text_area("Recomendaciones generales")

    # ── VIII. Hoja de vida (PDF) ────────────────────────────────────────────
    st.markdown('<div class="section-header">VIII. HOJA DE VIDA DEL CANDIDATO</div>',
                unsafe_allow_html=True)
    hv_pdfs = st.file_uploader(
        "Adjuntar archivos(PDF)",
        type=["pdf"],
        accept_multiple_files=True,
        help="Puedes adjuntar uno o varios archivos PDF. Tamaño máximo por archivo: 200 MB.",
        key="hoja_de_vida"
    )
    if hv_pdfs:
        for f in hv_pdfs:
            st.success(f"✅ Archivo cargado: **{f.name}** ({f.size / 1024:.1f} KB)")

    st.divider()

    # ── Botón de envío ────────────────────────────────────────────────────
    enviar = st.button("📤 Enviar Solicitud", use_container_width=True, type="primary")

    # ─────────────────────────────────────────────────────────────────────────
    # 11. PROCESAMIENTO AL ENVIAR
    # ─────────────────────────────────────────────────────────────────────────
    if enviar:
        errores = []
        if not Q8.strip():   errores.append("Razon social empresa es obligatorio.")
        if not Q10.strip():  errores.append("Nombre del AGR es obligatorio.")
        if not Q11.strip():  errores.append("Correo del AGR es obligatorio.")
        if not Q25:          errores.append("Debe seleccionar una ciudad.")
        if not Q26:          errores.append("Debe seleccionar al menos un dia de servicio.")

        if errores:
            for e in errores:
                st.error(e)
            st.stop()

        with st.spinner("Diligenciando formato Excel y enviando correo..."):
            id_sol    = generar_id_solicitud()
            fecha_hoy = datetime.date.today().strftime("%d/%m/%Y")

            datos = {
                "id_solicitud": id_sol,
                "Q2":  fecha_hoy,
                "Q6":  Q6,    "Q7":  Q7,    "Q8":  Q8,    "Q9":  Q9,
                "Q10": Q10,   "Q11": Q11,   "Q12": Q12,
                "Q13": Q13,   "Q14": Q14,
                "Q15": Q15,
                "Q20": Q20,   "Q21": Q21,
                "Q22": Q22,   "Q23": Q23,
                "Q24": Q24,   "Q25": Q25,
                "Q26": "; ".join(Q26),
                "Q27": Q27,   "Q28": Q28,   "Q29": Q29,
                "Q30": Q30,   "Q31": Q31,
                # Auxilios
                "Q32": Q32,   "Q33": Q33,   "Q34": Q34,
                "Q35": Q35,   "Q36": Q36,   "Q37": Q37,
                "Q38": Q38,   "Q39": Q39,   "Q40": Q40,
                "Q41": Q41,
                "Q42_texto": Q42_texto,
                "Q43_frec":  Q43_frec,
                "Q44_valor": Q44_valor,
                # Competencias
                "Q56": Q56,   "Q57": Q57,   "Q58": Q58,
                **comp_vals,
                "Q70": Q70,   "Q71_extra": Q71_extra,
                # EPPs
                **epp_vals,
                # Dotacion
                **dot_vals,
                "Q86": Q86,   "Q86_texto": Q86_texto,
                "Q87": Q87,   "Q88": Q88,
                "Q89": Q89,   "Q90": Q90,
                "Q91": Q91,
                "agr_data": agr_data if Q22 == "INTERDISCIPLINARIO" else [],
                "Q92": Q92,   "Q93": Q93,
            }

            logo_bytes_env = obtener_logo()

            plantilla_bytes = obtener_plantilla_excel()
            if plantilla_bytes is None:
                st.error("No se pudo obtener la plantilla Excel. Verifica que FORMATO.xlsx esté en el directorio.")
                st.stop()

            xlsx_bytes = diligenciar_formato_excel(datos, plantilla_bytes, logo_bytes_env)

            archivos_hv = []
            if hv_pdfs:
                for f in hv_pdfs:
                    f.seek(0)
                    archivos_hv.append({"bytes": f.read(), "nombre": f.name})

            exito       = enviar_correo(datos, xlsx_bytes, archivos_hv)
            exito_plano = enviar_archivo_plano(datos)

        asesora, clasificacion_info, dias_info = obtener_asesora_y_clasificacion(Q25)
        fecha_entrega = calcular_fecha_entrega(dias_info)
        nombre_base   = f"{id_sol} - {Q15} en {Q25}"

        if exito and exito_plano:
            st.success(f"✅ Solicitud **{id_sol}** enviada exitosamente a **{asesora}**.")
            st.balloons()
        elif exito and not exito_plano:
            st.warning("⚠️ La solicitud principal fue enviada, pero falló el envío del archivo plano.")
        elif (not exito) and exito_plano:
            st.warning("⚠️ Se envió el archivo plano, pero hubo un problema en el correo principal.")
        else:
            st.warning("⚠️ Solicitud procesada pero hubo problemas en los envíos de correo.")

        st.markdown(f"""
        <div class="info-box">
        Ciudad: <b>{Q25}</b> ({clasificacion_info}) &nbsp;|&nbsp;
        Tiempo de respuesta: <b>{dias_info} días hábiles</b> &nbsp;|&nbsp;
        Fecha tentativa de entrega: <b>{fecha_entrega}</b>
        </div>
        """, unsafe_allow_html=True)

        st.download_button(
            "⬇️ Descargar Excel Diligenciado",
            data=xlsx_bytes,
            file_name=f"{nombre_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


if __name__ == "__main__":
    main()
