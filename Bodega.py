import streamlit as st
import pandas as pd
from datetime import datetime
import openpyxl
import os
import smtplib
from email.message import EmailMessage


# ---------------------------------------
# INICIALIZACIÓN DEL TRIGGER DE RESETEO
if "reset_trigger" not in st.session_state:
    st.session_state.reset_trigger = 0

# Key dinámica basada en el trigger
reset_key = st.session_state.reset_trigger


# ---------------------------------------
@st.cache_data
def cargar_inventario():
    df = pd.read_excel('inventario.xlsx')
    return df


def actualizar_inventario_multiple(lista_productos):
    df = pd.read_excel('inventario.xlsx')
    resultado = []
    for producto in lista_productos:
        nombre_producto = producto['nombre']
        cantidad_solicitada = producto['cantidad']
        mascara = df['Nombre Producto'] == nombre_producto
        cantidad_actual = int(df.loc[mascara, 'Disponible'].values[0])
        nueva_cantidad = cantidad_actual - cantidad_solicitada
        df.loc[mascara, 'Disponible'] = nueva_cantidad
        resultado.append({
            'nombre': nombre_producto,
            'disponible_antes': cantidad_actual,
            'disponible_despues': nueva_cantidad
        })
    df.to_excel('inventario.xlsx', index=False)
    return resultado


def guardar_solicitud_excel(datos_personales, productos, plantilla_path="Plantilla.xlsx", salida_path="Solicitud_Completa.xlsx"):
    wb = openpyxl.load_workbook(plantilla_path)
    ws = wb.active
    inicio_filas = 9
    for idx, producto in enumerate(productos):
        fila = inicio_filas + idx
        ws[f"A{fila}"] = datos_personales["no_cc"]
        ws[f"B{fila}"] = datos_personales["nombre_trabajador"]
        ws[f"C{fila}"] = "ACTIVA "
        ws[f"D{fila}"] = datos_personales["convenio"]
        ws[f"E{fila}"] = datos_personales["cargo"]
        ws[f"F{fila}"] = datos_personales["correo"]
        ws[f"G{fila}"] = datos_personales["ciudad"]
        ws[f"H{fila}"] = datos_personales["direccion"]
        ws[f"I{fila}"] = datos_personales["telefono"]
        ws[f"J{fila}"] = producto["codigo"]
        ws[f"K{fila}"] = producto["nombre"]
        ws[f"L{fila}"] = producto["talla"]
        ws[f"M{fila}"] = producto["cantidad"]
        ws[f"N{fila}"] = "Incluido en tarifa"
        ws[f"O{fila}"] = "Incluido en tarifa"
        ws[f"P{fila}"] = ""
    wb.save(salida_path)


def enviar_email_con_adjuntos(destinatario, asunto, cuerpo, archivo_adjuntar, remitente, clave_app):
    msg = EmailMessage()
    msg["Subject"] = asunto
    msg["From"] = remitente
    msg["To"] = destinatario
    msg.set_content(cuerpo)
    with open(archivo_adjuntar, "rb") as f:
        file_data = f.read()
        file_name = os.path.basename(archivo_adjuntar)
    msg.add_attachment(file_data, maintype="application", subtype="octet-stream", filename=file_name)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(remitente, clave_app)
        smtp.send_message(msg)


# ---------------------------------------
# INICIALIZACIÓN DEL CARRITO CON KEY DINÁMICA
if f"carrito_productos_{reset_key}" not in st.session_state:
    st.session_state[f"carrito_productos_{reset_key}"] = []

# Variable local para el carrito
carrito = st.session_state[f"carrito_productos_{reset_key}"]


# ---------------------------------------
st.title("📦 Sistema de Solicitud de Dotación")

df = cargar_inventario()
df_filtrado = df[['Código','Nombre Producto', 'Disponible']].drop_duplicates()


st.markdown("### 📋 Datos del Trabajador")
col1, col2 = st.columns(2)
with col1:
    no_cc = st.text_input("NO CC", help="Número de cédula", key=f"no_cc_{reset_key}")
    nombre_trabajador = st.text_input("NOMBRE TRABAJADOR NOVASOFT", key=f"nombre_trabajador_{reset_key}")
    convenio = st.selectbox("CONVENIO", options=['Alkomprar', 'Alkosto', 'Cencosud', 'Claro Cavs', 'Claro Distribuidores', 'Claro TMK', 'Éxito', 'Falabella', 'Gestión', 'Ktronix', 'Mayoristas', 'Movistar', 'Movistar Phoenix', 'Supernumerario', 'Nariño', 'Olimpica', 'Panamericana', 'Supernumerario Claro', 'Tigo', 'Wom', 'Xperts Mixtos'], key=f"convenio_{reset_key}")
    cargo = st.selectbox("CARGO", options=['EJECUTIVO DE CUENTA', 'EJECUTIVO DE CUENTA B2B', 'EJECUTIVO DE CUENTA JR', 'ENTRENADOR', 'ENTRENADOR INTEGRAL', 'ENTRENADOR JUNIOR', 'ENTRENADOR MASTER', 'ESPECIALISTA DE EXPERIENCIA', 'JEFE DE ZONA', 'MERCHANDISER (EXPERT)', 'SUPERNUMERARIO ADECCO'], key=f"cargo_{reset_key}")
with col2:
    correo = st.text_input("CORREO ELECTRONICO NOVASOFT", key=f"correo_{reset_key}")
    ciudad = st.selectbox("CIUDAD DE ENTREGA O LABOR NOVASOFT", options=['Aguachica', 'Apartadó', 'Arauca', 'Arauquita', 'Armenia', 'Barrancabermeja', 'Barranquilla', 'Bello', 'Bogotá', 'Bucaramanga', 'Buenaventura', 'Cali', 'Cartagena', 'Cartago', 'Caucasia', 'Chía', 'Cúcuta', 'Duitama', 'Facatativá', 'Florencia', 'Fusagasugá', 'Girardot', 'Granada', 'Ibagué', 'Ipiales', 'Itagüí', 'La Dorada', 'Maicao', 'Manizales', 'Medellín', 'Montería', 'Mosquera', 'Neiva', 'Palmira', 'Pasto', 'Pereira', 'Piedecuesta', 'Popayán', 'Quibdó', 'Riohacha', 'Rionegro', 'San Gil', 'Santa Marta', 'Santander De Quilichao', 'Sincelejo', 'Soacha', 'Sogamoso', 'Tuluá', 'Tunja', 'Túquerres', 'Valledupar', 'Villavicencio', 'Yopal', 'Zipaquirá'], key=f"ciudad_{reset_key}")
    telefono = st.text_input("TELÉFONO DE CONTACTO", key=f"telefono_{reset_key}")
direccion = st.text_input("DIRECCIÓN ENTREGA (INCLUIR TORRE, APTO, CONJ, CASA)", key=f"direccion_{reset_key}")


st.markdown("---")
st.markdown("### Agregar Productos")
nombre_producto = st.selectbox(
    "NOMBRE DEL PRODUCTO",
    options=df_filtrado['Nombre Producto'].tolist(),
    index=None,
    placeholder="Busca y selecciona un producto...",
    help="Escribe para buscar el producto",
    key=f"nombre_producto_{reset_key}"
)

if nombre_producto is not None:
    disponible = int(df_filtrado.set_index('Nombre Producto').loc[nombre_producto, 'Disponible'])
    if disponible > 0:
        st.success(f"✅ **Cantidad disponible:** {disponible} unidades")
    else:
        st.error(f"❌ **Sin stock disponible** - Este producto no tiene unidades disponibles")
else:
    disponible = 0
    st.info("⬆️ Selecciona un producto para ver la disponibilidad")


if nombre_producto is not None and disponible > 0:
    col_t, col_c = st.columns(2)
    with col_t:
        talla_options = ['3', '4', '6', '8', '10', '12', '14', '16', '18', '20', '28', '30', '32', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', 'L', 'M', 'S', 'XL', 'XS', 'XXL']
        talla_temp = st.selectbox("TALLA", options=talla_options, key=f"talla_temp_{reset_key}", placeholder="Selecciona una talla...")
    with col_c:
        cantidad_temp = st.number_input(
            "CANTIDAD REQUERIDA",
            min_value=1,
            max_value=disponible,
            value=1,
            key=f"cantidad_temp_{reset_key}",
            help=f"Máximo disponible: {disponible}"
        )
    if st.button("➕ Agregar al Carrito", type="secondary", key=f"agregar_{reset_key}"):
        if not talla_temp:
            st.error("❌ Debes ingresar la talla")
        else:
            codigo = df_filtrado.set_index('Nombre Producto').loc[nombre_producto, 'Código']
            carrito.append({
                'nombre': nombre_producto,
                'codigo': codigo,
                'talla': talla_temp,
                'cantidad': cantidad_temp,
                'disponible': disponible
            })
            st.success(f"✅ Producto agregado al carrito: {nombre_producto} (Talla: {talla_temp}, Cantidad: {cantidad_temp})")
            st.rerun()


st.markdown("---")
st.markdown("### Productos seleccionados")
if len(carrito) > 0:
    for idx, producto in enumerate(carrito):
        col1, col2 = st.columns([5, 1])
        with col1:
            st.write(f"**{producto['nombre']}** - Talla: {producto['talla']}, Cantidad: {producto['cantidad']}")
        with col2:
            if st.button("🗑️ Eliminar", key=f"eliminar_{idx}_{reset_key}"):
                carrito.pop(idx)
                st.rerun()
    if st.button("🗑️ Vaciar productos", type="secondary", key=f"vaciar_{reset_key}"):
        st.session_state[f"carrito_productos_{reset_key}"] = []
        st.rerun()
else:
    st.info("No ha seleccionado ningún producto. Agrega productos antes de enviar la solicitud.")


st.markdown("---")
enviar = st.button("📤 Enviar Solicitud Completa", type="primary", key=f"enviar_{reset_key}")


if enviar:
    errores = []
    if len(carrito) == 0:
        errores.append("Debes agregar al menos un producto al carrito")
    if not no_cc:
        errores.append("Ingresa el número de CC")
    if not nombre_trabajador:
        errores.append("Ingresa el nombre del trabajador")
    if not correo:
        errores.append("Ingresa el correo electrónico")
    if not telefono:
        errores.append("Ingresa el teléfono de contacto")
    if not direccion:
        errores.append("Ingresa la dirección de entrega")
    if errores:
        st.error("❌ Por favor completa los siguientes campos:")
        for error in errores:
            st.write(f"- {error}")
    else:
        try:
            resultado_actualizacion = actualizar_inventario_multiple(carrito)
            datos_personales = {
                "no_cc": no_cc,
                "nombre_trabajador": nombre_trabajador,
                "convenio": convenio,
                "cargo": cargo,
                "correo": correo,
                "ciudad": ciudad,
                "direccion": direccion,
                "telefono": telefono
            }
            guardar_solicitud_excel(datos_personales, carrito)
            st.cache_data.clear()
            st.success(f"✅ ¡Solicitud enviada correctamente! Se procesaron {len(carrito)} productos.")
            with st.expander("📄 Ver detalles de la actualización de inventario"):
                for res in resultado_actualizacion:
                    st.write(f"**{res['nombre']}:** {res['disponible_antes']} → {res['disponible_despues']} unidades")
            
            # Enviar email
            remitente = "dotacionmotorolaadecco@gmail.com"
            clave_app = "jajx eggv rizf sndd"
            try:
                enviar_email_con_adjuntos(
                    destinatario="Manuel.Pimentel@adecco.com",
                    asunto="Nueva solicitud de dotación",
                    cuerpo=" Se adjunta la solicitud en formato Excel.",
                    archivo_adjuntar="Solicitud_Completa.xlsx",
                    remitente=remitente,
                    clave_app=clave_app
                )
                st.success("✅ Se envió la solicitud por correo electrónico correctamente.")
            except Exception as err_mail:
                st.error(f"❌ Error al enviar el correo: {str(err_mail)}")
            
            st.info("🔄 Usa el botón 'Resetear Aplicación Completa' para hacer una nueva solicitud")
        except Exception as e:
            st.error(f"❌ Error al procesar la solicitud: {str(e)}")


with st.expander("📊 Ver inventario completo"):
    st.dataframe(df_filtrado, use_container_width=True)


with st.expander("🔄 Cargar nuevo inventario (Excel)"):
    archivo_nuevo = st.file_uploader(
        "Selecciona el archivo Excel (.xlsx) para cargar un nuevo inventario y reemplazar el existente:",
        type=["xlsx"],
        key=f"uploader_{reset_key}"
    )
    if archivo_nuevo is not None:
        try:
            df_nuevo = pd.read_excel(archivo_nuevo)
            df_nuevo.to_excel("inventario.xlsx", index=False)
            st.success("✅ Inventario reemplazado exitosamente con los datos del archivo cargado.")
            st.cache_data.clear()
            st.info("🔄 Recarga la página para ver el inventario actualizado.")
        except Exception as e:
            st.error(f"❌ Error al cargar el archivo: {str(e)}")


st.markdown("---")
if st.button("🔄 Crear nueva solicitud", key=f"reset_button_{reset_key}"):
    # Incrementar el trigger para forzar nueva inicialización
    st.session_state.reset_trigger += 1
    # Limpiar caché
    st.cache_data.clear()
    # Recargar la aplicación
    st.rerun()
